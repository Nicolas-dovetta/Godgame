"""LIFE — 3-scenario financial model generator.
Writes computed VALUES (each number verified against the transparent logic below).
Edit the DRIVERS block and re-run to flex any scenario. No external deps beyond openpyxl.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------- DRIVERS
DRIVERS = [
    # (label, bear, base, bull, fmt_key, note)
    ("New registrations — month 1",         700,   1800,  3500,  "num",  "Monthly new sign-ups at launch"),
    ("Monthly growth in new registrations", 0.065, 0.10,  0.14,  "pct",  "Compounds MoM; the chronicle-is-the-ad loop"),
    ("Free monthly churn",                  0.12,  0.09,  0.07,  "pct",  "Free base attrition / month"),
    ("Monthly free->paid conversion",       0.012, 0.020, 0.030, "pct",  "% of free base upgrading each month"),
    ("Paid monthly churn",                  0.08,  0.05,  0.035, "pct",  "Subscriber attrition / month"),
    ("Free COGS $/user/mo",                 0.20,  0.20,  0.22,  "usd2", "1 turn/night on Grok 4.1 Fast + moderation"),
    ("Paid COGS $/user/mo",                 3.20,  3.20,  3.40,  "usd2", "~2.5 turns/day avg on Grok 4.5 + moderation"),
    ("Subscription price $/mo",             12,    12,    12,    "usd",  "Monthly - Deity tier"),
    ("Payment fees % of revenue",           0.035, 0.035, 0.035, "pct",  "Stripe ~2.9% + $0.30 blended"),
    ("Whale share of payers",               0.08,  0.12,  0.18,  "pct",  "Heavy a-la-carte spenders"),
    ("Whale a-la-carte $/mo",               10,    18,    28,    "usd",  "Grace + experience packs, per whale"),
    ("Regular payer a-la-carte $/mo",       0.75,  1.25,  2.00,  "usd2", "Non-whale payers, per month"),
    ("Free-user impulse a-la-carte $/mo",   0.03,  0.06,  0.12,  "usd2", "Occasional pack buy by free users"),
    ("A-la-carte COGS % of a-la-carte rev", 0.45,  0.45,  0.45,  "pct",  "Grace = turns priced ~2x their compute cost"),
    ("xAI free credits $/mo (COGS offset)", 175,   175,   175,   "usd",  "Data-sharing program; offsets early COGS"),
    ("Fixed opex $/mo — Year 1",            3000,  5000,  8000,  "usd",  "Founder draw, tools, part-time T&S"),
    ("Fixed opex $/mo — Year 2",            5000,  15000, 35000, "usd",  "+ hires as it scales"),
    ("Fixed opex $/mo — Year 3",            8000,  30000, 80000, "usd",  "Team, T&S, support, infra"),
    ("One-time legal review (month 4) $",   1500,  1500,  1500,  "usd",  "ToS/privacy/data-consent lawyer review"),
    ("Starting cash $",                     0,     0,     0,     "usd",  "Bootstrapped, no external funding"),
]
KEYS = ["regs","grow","fchurn","conv","pchurn","fcogs","pcogs","price","fee","wshare",
        "wspend","rspend","fspend","acogs","credit","ox1","ox2","ox3","legal","cash0"]

def params(col):  # col: 0 bear, 1 base, 2 bull
    d = {KEYS[i]: DRIVERS[i][1+col] for i in range(len(KEYS))}
    d["ox"] = (d["ox1"], d["ox2"], d["ox3"])
    return d

def run(p):
    rows = []
    for m in range(1, 37):
        prev = rows[-1] if rows else None
        regs = p["regs"] if m == 1 else prev["regs"]*(1+p["grow"])
        conv = 0.0 if m == 1 else p["conv"]*prev["free"]
        free = (regs-conv) if m == 1 else prev["free"]*(1-p["fchurn"])+regs-conv
        paid = conv if m == 1 else prev["paid"]*(1-p["pchurn"])+conv
        whales = p["wshare"]*paid
        sub = paid*p["price"]*(1-p["fee"])
        ac_gross = whales*p["wspend"] + (paid-whales)*p["rspend"] + free*p["fspend"]
        ac = ac_gross*(1-p["fee"])
        rev = sub+ac
        fcogs, pcogs, mcogs = free*p["fcogs"], paid*p["pcogs"], ac_gross*p["acogs"]
        credit = min(p["credit"], fcogs+pcogs+mcogs)
        cogs = fcogs+pcogs+mcogs-credit
        gp = rev-cogs
        oxb = p["ox"][0] if m <= 12 else (p["ox"][1] if m <= 24 else p["ox"][2])
        ox = oxb + (p["legal"] if m == 4 else 0)
        cf = gp-ox
        cum = (p["cash0"]+cf) if not rows else prev["cum"]+cf
        rows.append(dict(m=m, year=(m-1)//12+1, regs=regs, free=free, conv=conv, paid=paid,
                         whales=whales, sub=sub, ac=ac, rev=rev, fcogs=fcogs, pcogs=pcogs,
                         mcogs=mcogs, credit=credit, cogs=cogs, gp=gp, ox=ox, cf=cf, cum=cum))
    return rows

# ----------------------------------------------------------------- STYLES
def F(**k): return Font(name="Arial", size=k.pop("size",10), **k)
TITLE = F(size=14, bold=True)
SUB   = Font(name="Arial", size=9, italic=True, color="666666")
HDR   = F(bold=True, color="FFFFFF")
BOLD  = F(bold=True)
BLUE  = F(color="0000FF")
BLK   = F()
HDRFILL = PatternFill("solid", fgColor="1B1B28")
YEL   = PatternFill("solid", fgColor="FFF7CC")
GREY  = PatternFill("solid", fgColor="F2F0F6")
FMT = {"num":'#,##0;(#,##0);-', "usd":'$#,##0;($#,##0);-',
       "usd2":'$#,##0.00;($#,##0.00);-', "pct":'0.0%'}

wb = openpyxl.Workbook()

# ----------------------------------------------------------------- SUMMARY
s = wb.active; s.title = "Summary"; s.sheet_view.showGridLines = False
s["A1"] = "LIFE — 3-Scenario Financials"; s["A1"].font = TITLE
s["A2"] = ("Grok-only engine · mature (non-explicit) · $0-capex · PWA-first · 1 turn/night free, 4/day paid. "
           "$ unless noted. Illustrative model, not a forecast — values computed from the Assumptions tab.")
s["A2"].font = SUB

scen_names = ["Bear", "Base", "Bull"]
data = {n: run(params(i)) for i, n in enumerate(scen_names)}

def yr(rows, y):
    seg = [r for r in rows if r["year"] == y]
    tot = lambda k: sum(r[k] for r in seg)
    return dict(regs=tot("regs"), rev=tot("rev"), cogs=tot("cogs"), gp=tot("gp"),
                ox=tot("ox"), cf=tot("cf"), free=seg[-1]["free"], paid=seg[-1]["paid"],
                whales=seg[-1]["whales"], cash=seg[-1]["cum"],
                gm=(tot("gp")/tot("rev") if tot("rev") else 0))

metrics = [
    ("New registrations",           "regs", "num"),
    ("Ending free users",           "free", "num"),
    ("Ending paid users",           "paid", "num"),
    ("Ending whale spenders",       "whales","num"),
    ("Revenue",                     "rev",  "usd"),
    ("COGS",                        "cogs", "usd"),
    ("Gross profit",                "gp",   "usd"),
    ("Gross margin %",              "gm",   "pct"),
    ("Operating expenses",          "ox",   "usd"),
    ("Operating cash flow",         "cf",   "usd"),
    ("Ending cash position",        "cash", "usd"),
]
row = 4
for n in scen_names:
    rows = data[n]
    s.cell(row=row, column=1, value=f"{n.upper()} scenario")
    for j, h in enumerate(["", "Year 1", "Year 2", "Year 3", "3-yr total"]):
        c = s.cell(row=row, column=1+j, value=(f"{n.upper()} scenario" if j == 0 else h))
        c.font = HDR; c.fill = HDRFILL; c.alignment = Alignment(horizontal="center" if j else "left")
    row += 1
    yrs = {y: yr(rows, y) for y in (1, 2, 3)}
    for label, key, fk in metrics:
        s.cell(row=row, column=1, value=label).font = BLK
        for yi, y in enumerate((1, 2, 3)):
            c = s.cell(row=row, column=2+yi, value=round(yrs[y][key], 4 if fk == "pct" else 0))
            c.number_format = FMT[fk]; c.font = BLK
        # 3-yr total / ending
        if key in ("free", "paid", "whales", "cash"):
            tv = yrs[3][key]
        elif key == "gm":
            tr = sum(r["rev"] for r in rows); tv = (sum(r["gp"] for r in rows)/tr) if tr else 0
        else:
            tv = sum(r[key] for r in rows)
        c = s.cell(row=row, column=5, value=round(tv, 4 if fk == "pct" else 0))
        c.number_format = FMT[fk]; c.font = BOLD
        if label in ("Revenue", "Operating cash flow", "Ending paid users"):
            for cc in range(1, 6): s.cell(row=row, column=cc).fill = GREY
        row += 1
    row += 1

# headline block
s.cell(row=row, column=1, value="3-YEAR HEADLINE")
for j, h in enumerate(["3-year headline", "Bear", "Base", "Bull"]):
    c = s.cell(row=row, column=1+j, value=h); c.font = HDR; c.fill = HDRFILL
    c.alignment = Alignment(horizontal="center" if j else "left")
row += 1
def first_cf_pos(rows): return next((r["m"] for r in rows if r["cf"] > 0), None)
headline = [
    ("3-yr revenue",                     lambda r: sum(x["rev"] for x in r), "usd"),
    ("3-yr operating cash flow",         lambda r: sum(x["cf"] for x in r),  "usd"),
    ("Funding trough (min cash)",        lambda r: min(x["cum"] for x in r), "usd"),
    ("First cash-flow-positive month",   lambda r: first_cf_pos(r),          "num"),
    ("Ending paid users (M36)",          lambda r: r[-1]["paid"],            "num"),
    ("M36 revenue run-rate (annualized)",lambda r: r[-1]["rev"]*12,          "usd"),
    ("Ending cash (M36)",                lambda r: r[-1]["cum"],             "usd"),
]
for label, fn, fk in headline:
    s.cell(row=row, column=1, value=label).font = BLK
    for j, n in enumerate(scen_names):
        c = s.cell(row=row, column=2+j, value=round(fn(data[n]) or 0, 0)); c.number_format = FMT[fk]; c.font = BOLD
    row += 1
s.column_dimensions["A"].width = 32
for col in "BCDE": s.column_dimensions[col].width = 15

# ----------------------------------------------------------------- ASSUMPTIONS
a = wb.create_sheet("Assumptions"); a.sheet_view.showGridLines = False
a["A1"] = "Assumptions (drivers)"; a["A1"].font = TITLE
a["A2"] = "Blue = the levers. This workbook is a computed snapshot — edit build_model.py's DRIVERS and re-run to change scenarios."
a["A2"].font = SUB
for j, h in enumerate(["Driver", "Bear", "Base", "Bull", "Notes"]):
    c = a.cell(row=4, column=1+j, value=h); c.font = HDR; c.fill = HDRFILL
    c.alignment = Alignment(horizontal="center" if 0 < j < 4 else "left")
for i, (label, b, ba, bu, fk, note) in enumerate(DRIVERS):
    r = 5+i
    a.cell(row=r, column=1, value=label).font = BLK
    for j, v in enumerate((b, ba, bu)):
        c = a.cell(row=r, column=2+j, value=v); c.font = BLUE; c.number_format = FMT[fk]
        c.fill = YEL; c.alignment = Alignment(horizontal="center")
    a.cell(row=r, column=5, value=note).font = SUB
a.column_dimensions["A"].width = 34
for col in "BCD": a.column_dimensions[col].width = 12
a.column_dimensions["E"].width = 48

# ----------------------------------------------------------------- MODEL SHEETS
cols = [("m","Month","num"),("year","Year","num"),("regs","New regs","num"),
        ("free","Free base","num"),("conv","Converts","num"),("paid","Paid base","num"),
        ("whales","Whales","num"),("sub","Sub rev (net)","usd"),("ac","A-la-carte (net)","usd"),
        ("rev","Total revenue","usd"),("fcogs","Free COGS","usd"),("pcogs","Paid COGS","usd"),
        ("mcogs","A-la-carte COGS","usd"),("credit","xAI credit","usd"),("cogs","Total COGS","usd"),
        ("gp","Gross profit","usd"),("ox","Opex","usd"),("cf","Op cash flow","usd"),("cum","Cumulative cash","usd")]
for n in scen_names:
    ws = wb.create_sheet(f"Model — {n}"); ws.sheet_view.showGridLines = False
    ws["A1"] = f"{n} scenario — 36-month model"; ws["A1"].font = TITLE
    for j, (_, h, _) in enumerate(cols):
        c = ws.cell(row=3, column=1+j, value=h); c.font = HDR; c.fill = HDRFILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for r in data[n]:
        rr = 3+r["m"]
        for j, (key, _, fk) in enumerate(cols):
            c = ws.cell(row=rr, column=1+j, value=round(r[key], 0) if fk != "pct" else r[key])
            c.number_format = FMT[fk] if key not in ("m","year") else "0"; c.font = BLK
        if r["m"] % 12 == 0:
            for j in range(len(cols)): ws.cell(row=rr, column=1+j).fill = GREY
    ws.column_dimensions["A"].width = 6; ws.column_dimensions["B"].width = 6
    for j in range(3, 20): ws.column_dimensions[get_column_letter(j)].width = 13
    ws.freeze_panes = "C4"

out = "/tmp/claude-0/-home-user-Godgame/b20ef510-d840-51cd-99f3-3a96122a0574/scratchpad/LIFE-financial-model.xlsx"
wb.save(out)
print("saved", out, "sheets:", wb.sheetnames)
